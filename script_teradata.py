import os 
import re
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
folder = 'logs'
def extract_columns_from_log(log_file_path):
    # Vérifier si le fichier existe
    if not os.path.isfile(log_file_path):
        print(f"This file {log_file_path} doesn't exist.")
        return []
    # Lire le fichier log
    with open(log_file_path, 'r', encoding='utf-8') as file:
        log_lines = file.readlines()
        columns = []
        inside_teradata = False
        type_of_data=''
        data = []
        name_of_table=''
        complexe_query_teradata=False
        liste_of_table=[]
        date = None
        result=''
        job=str(log_file_path)
        i=0
        pattern = re.compile(r'(\d+)\s+rows.*?(\d+)\s+columns', re.IGNORECASE)
        for line in log_lines:
            if 'SELECT' in line and 'FROM CONNECTION TO TERADATA' in line:
                inside_teradata = True
                type_of_data = 'Teradata'
                print(f'line is inside teradata,{i}')
            if 'DISCONNECT FROM TERADATA' in line :
                inside_teradata=False
                type_of_data = 'SAS'
                print(f'line is outside teradata,{i}')
            if 'FROM' in line and  inside_teradata and not 'SELECT' in line:
                patern = re.compile(r'FROM\s+(\S+)',re.IGNORECASE)
                resultat = patern.search(line)
                if resultat : 
                    name_of_table = resultat.group(1)
            if  inside_teradata and 'NOTE: ' in line :
                match = pattern.search(line)
                if match:
                    rows = int(match.group(1))
                    columns = int(match.group(2))
                    print(rows,columns,resultat.group(1),type_of_data)
                    data.append([name_of_table,rows,columns,type_of_data,date,'read',job])
            if 'The SAS System ' in line and '2024' in line and date == None :
                date_pattern = r"The SAS System\s+\w+, (\w+ \d{1,2}, \d{4})"
                match = re.search(date_pattern, line)
                if match:
                        date_str = match.group(1)
                        # Convertir la date en objet datetime
                        date_obj = datetime.strptime(date_str, "%B %d, %Y")
                        # Formatter la date en dd-mm-yyyy
                        formatted_date = date_obj.strftime("%d-%m-%Y")
                        date = formatted_date
            else:
                i=i+1
        colonnes = ['Table source', 'Number of rows','Number of columns' ,'Database','Extraction date','Action','JOB']
        df = pd.DataFrame(data, columns=colonnes)
        wb = load_workbook('data_file.xlsx')
        ws = wb.active
        row_num = ws.max_row + 1
        for index, row in df.iterrows():
            for col_num, value in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num, value=value)
            row_num += 1
        wb.save('data_file.xlsx')
        print(df)
        print(f"We analyzed {i} rows")
dossier = Path('logs')
extract_columns_from_log('test_oracle.log')
for fichier in dossier.iterdir():#Step 1 extract data from the excel file 
        # Vérifier si c'est un fichier (et non un dossier)
        print('The script has started')
        if fichier.is_file():
            print(fichier)
            extract_columns_from_log(fichier)
        print('I finished with the script')